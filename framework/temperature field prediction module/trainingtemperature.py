import numpy as np
import torch
from torch import nn
import sys
from openpyxl import Workbook
from torch.utils.data import Dataset, DataLoader
sys.path.append('./')

# Define mini-batch
class CustomDataset(Dataset):
    def __init__(self, inputs1, inputs2, labels):
        self.len=inputs1.shape[0]
        self.inputs1=torch.from_numpy(inputs1)
        self.inputs2=torch.from_numpy(inputs2)
        self.labels=torch.from_numpy(labels)

    def __len__(self):
        return self.len

    def __getitem__(self, index):
        return self.inputs1[index], self.inputs2[index], self.labels[index]


# Define CNN Neural Networks
class CNN(nn.Module):
    def __init__(self):
        super().__init__()
        self.dropout = nn.Dropout(0.3)
        self.conv1 = nn.Conv2d(1, 75, 3)
        self.pool = nn.MaxPool2d(3, 2)
        self.conv2 = nn.Conv2d(75, 150, 2)
        self.linear = nn.Linear(150, 150)

    def forward(self, x):
        x=self.dropout(x)
        x = x.float()
        x = self.pool(torch.relu(self.conv1(x)))
        x = self.pool(torch.relu(self.conv2(x)))
        x = x.view(x.size(0), -1)
        x = self.linear(x)
        return x

# Define CNN-LSTM Neural Networks
class CNNLSTM(nn.Module):
    def __init__(self):
        super().__init__()
        self.cnn = CNN()
        self.lstm = nn.LSTM(150, 200, 2, batch_first=True, dropout=0.3)
        self.linear1 = nn.Linear(200, 195)
        self.linear2 = nn.Linear(200, 200)
        self.linear3 = nn.Linear(200, 200)
        self.linear4 = nn.Linear(200, 121)
        self.ReLU=nn.ReLU()

    def forward(self, x1, x2):
        batch_size, sequence = x1.shape[0], x1.shape[1]
        data_cnn_x1 = torch.zeros(batch_size, sequence, 150).to(x1.device)

        for i in range(sequence):
            data_cnn = x1[:, i, :]
            data_cnn = data_cnn.reshape(-1, 1, 11, 11)
            data_cnn = self.cnn(data_cnn)
            data_cnn = data_cnn.reshape(-1, 1, 150)
            data_cnn_x1[:, i, :] = data_cnn.squeeze(1)

        x, _ = self.lstm(data_cnn_x1)  # x1 is input, size (batch, seq_len, input_size) 对时序数据进行lstm

        data_dense = x[:, -1, :]  # 取最后一个时间点的结果
        data_dense = data_dense.unsqueeze(1)
        data_dense = self.linear1(data_dense)
        data_dense = self.ReLU(data_dense)
        data_dense = torch.cat([data_dense, x2], dim=2)
        data_dense = data_dense.float()
        data_dense = self.linear2(data_dense)
        data_dense = self.ReLU(data_dense)
        data_dense = self.linear3(data_dense)
        data_dense = self.ReLU(data_dense)
        data_dense = self.linear4(data_dense)
        return data_dense

if __name__ == '__main__':
    train_data_ratio = 0.8  # Choose 80% of the data for training
    max_epochs=1000  #训练循环次数
    batch_size=64
    batch_size_test=64
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    current_device = torch.cuda.current_device()

# ================================ 数据导入和处理
    data_x2_3 = np.load("data_x2_3.npy")
    data_y_3 = np.load("data_y_3.npy")
    data_x2= data_x2_3[:, ::10, :]
    data_y=data_y_3[:, ::10, :]

# ==========================================数据库处理
    np.random.seed(10)
    np.random.shuffle(data_x2)
    np.random.seed(10)
    np.random.shuffle(data_y)    # 将数据组打乱，为测试集和试验集的选取做准备

    total_batch_size, seq_len = data_y.shape[0],data_y.shape[1]
    train_data_len = int(total_batch_size * train_data_ratio)
    data_x1=np.full((total_batch_size,seq_len,121),0)  # 经过归一化后常温20度为0

    train_x1 = data_x1[0:train_data_len]  # 取train_data_len个数据做训练集
    train_x2 = data_x2[0:train_data_len]
    train_y = data_y[0:train_data_len]

    # train数据集需要minibatch，test数据集不需要
    Mydataset = CustomDataset(train_x1, train_x2, train_y)
    train_loader = DataLoader(dataset=Mydataset, batch_size=batch_size, shuffle=False, num_workers=1)

    test_x1 = data_x1[train_data_len:]  # 取剩下数据做测试集
    test_x2 = data_x2[train_data_len:]
    test_y = data_y[train_data_len:]
    
    Mydataset_test = CustomDataset(test_x1, test_x2, test_y)
    test_loader = DataLoader(dataset=Mydataset_test, batch_size=batch_size_test, shuffle=False, num_workers=0)

# ================================计算cnnlstm模型
    cnnlstm_model = CNNLSTM().to(device)
    cnnlstm_model=cnnlstm_model.float()
    print('model:', cnnlstm_model)
    print('model.parameters:', cnnlstm_model.parameters)

    criterion = nn.MSELoss()
    optimizer = torch.optim.Adam(cnnlstm_model.parameters(), lr=1e-4)

    prev_loss = 1000
    loss_value = []
    test_loss_value = []
    for epoch in range(max_epochs):
        for i, data in enumerate(train_loader, 0):
            train_x1_tensor, train_x2_tensor, train_y_tensor = data
            train_x1_tensor=train_x1_tensor.float().to(device)
            train_x2_tensor=train_x2_tensor.float().to(device)
            train_y_tensor=train_y_tensor.float().to(device)
            print('第',i+1,'个batch')
            for j in range(seq_len):
                train_x2_one_tensor=train_x2_tensor[:,j,:]  
                train_x2_one_tensor=train_x2_one_tensor.reshape(-1,1,5)
                train_y_one_tensor=train_y_tensor[:,j,:]  
                train_y_one_tensor=train_y_one_tensor.reshape(-1,1,121)
                output = cnnlstm_model(train_x1_tensor, train_x2_one_tensor)
                loss = criterion(output, train_y_one_tensor)
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
                train_x1_tensor = torch.cat([train_x1_tensor[:, 1:seq_len, :], output.detach()], dim=1)
                del train_x2_one_tensor
                del train_y_one_tensor
                del output
                torch.cuda.empty_cache()
                loss_value_w=np.expand_dims(loss.item(),axis=0)
                loss_value.extend(loss_value_w)
        trainbatch_geshu=i+1
        trainseq_len=j+1
            
        for k, data_test in enumerate(test_loader, 0):
            test_x1_tensor, test_x2_tensor, test_y_tensor = data_test
            test_x1_tensor=test_x1_tensor.float().to(device)
            test_x2_tensor=test_x2_tensor.float().to(device)
            test_y_tensor=test_y_tensor.float().to(device)
            for m in range(seq_len):
                test_x2_one_tensor=test_x2_tensor[:,m,:]  
                test_x2_one_tensor=test_x2_one_tensor.reshape(-1,1,5)
                test_y_one_tensor=test_y_tensor[:,m,:]
                test_y_one_tensor=test_y_one_tensor.reshape(-1,1,121)
                output_test = cnnlstm_model(test_x1_tensor, test_x2_one_tensor)
                loss_test = criterion(output_test, test_y_one_tensor)
                test_x1_tensor = torch.cat([test_x1_tensor[:, 1:seq_len, :], output_test.detach()], dim=1)
                del test_x2_one_tensor
                del test_y_one_tensor
                del output_test
                torch.cuda.empty_cache()
                test_loss_value_w=np.expand_dims(loss_test.item(),axis=0)
                test_loss_value.extend(test_loss_value_w)

        if loss < prev_loss:
            torch.save(cnnlstm_model.state_dict(), 'cnnlstm4.pt')  # save model parameters to files
            prev_loss = loss

        if loss.item() < 1e-7:
            print('Epoch [{}/{}], Loss: {:.5f}'.format(epoch + 1, max_epochs, loss.item()))
            print("The loss value is reached")
            break
        elif (epoch + 1) % 1 == 0:
            print('Epoch: [{}/{}], Loss:{:.5f}'.format(epoch + 1, max_epochs, loss.item()))
    trainbatch_geshu=i+1
    trainseq_len=j+1  
    testbatch_geshu=k+1
    testseq_len=m+1

# ===============================输出loss
    loss_value=np.array(loss_value).reshape(-1,1)
    loss_geshu=loss_value.shape[0]
    test_loss_value=np.array(test_loss_value).reshape(-1,1)
    test_loss_geshu=test_loss_value.shape[0]
    wb = Workbook()
    ws1 = wb.active
    ws2=wb.create_sheet()
    ws1.title = "train loss"
    row=0
    for i in range(1, loss_geshu+1,trainbatch_geshu*trainseq_len):
        row = row + 1
        for j in range(1,trainbatch_geshu*trainseq_len+1):
            ws1.cell(row=row, column=j).value = loss_value[i + j - 2, 0]
    ws2.title = "test loss"
    row=0
    for i in range(1, test_loss_geshu+1,testbatch_geshu*testseq_len):
        row = row + 1
        for j in range(1,testbatch_geshu*testseq_len+1):
            ws2.cell(row=row, column=j).value = test_loss_value[i + j - 2, 0]
    wb.save('loss_4.xlsx')