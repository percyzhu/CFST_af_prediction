import numpy as np
import torch
from torch import nn
import sys
from openpyxl import Workbook
from torch.utils.data import Dataset, DataLoader
from torch.nn.utils import clip_grad_norm_
sys.path.append('./')

# Define mini-batch
class CustomDataset(Dataset):
    def __init__(self, inputs1, inputs2, inputs3, inputs4):
        self.len=inputs1.shape[0]
        self.inputs1=torch.from_numpy(inputs1)
        self.inputs2=torch.from_numpy(inputs2)
        self.inputs3=torch.from_numpy(inputs3)
        self.inputs4=torch.from_numpy(inputs4)

    def __len__(self):
        return self.len

    def __getitem__(self, index):
        return self.inputs1[index], self.inputs2[index], self.inputs3[index], self.inputs4[index]

# Define PELSTM Neural Networks
class PELSTM(nn.Module):
    def __init__(self):
        super().__init__()
        self.linear1 = nn.Linear(125, 200)
        self.linear2 = nn.Linear(200, 200)
        self.linear3 = nn.Linear(200, 196)
        self.lstm = nn.LSTM(200, 10, 2, batch_first=True, dropout=0.3)
        self.linear4 = nn.Linear(10,1)
        self.ReLU=nn.ReLU()

    def forward(self, x1, x2):
        batch_size, sequence = x1.shape[0], x1.shape[1]
        data_x = torch.zeros(batch_size, sequence, 200).to(x1.device)
        
        input_x=torch.cat([x1, x2], dim=2)
        input_x=self.linear1(input_x)
        input_x = self.ReLU(input_x)
        input_x=self.linear2(input_x)
        input_x = self.ReLU(input_x)
        input_x=self.linear3(input_x)
        input_x= self.ReLU(input_x)
        data_x =torch.cat([input_x, x2], dim=2)

        x, _ = self.lstm(data_x)  # 对时序数据进行lstm
        x=x[:, -1, :]  # 取最后一个时间点的轴力比
        x=self.linear4(x)
        return x
 
# 数据损失
def data_loss(model, x1, x2, x3):
    p = model(x1, x2)  # lstm计算得到的轴力比
    y = x3[:, -1, :]  # 取最后一个时间点的轴力比
    return torch.mean((p - y) ** 2)    

# 物理损失
def physics_loss(model, x1, x2, x3, x2_real):   
    p = model(x1, x2)  # lstm计算得到的轴力比
    p_former = x3[:, 2, :]  # 取前一个时间点的轴力比
    s = x2_real[:, 3, 3]  # 取最后一个时间点的变形速率
    x = x2_real[:, 3, 2]  # 取最后一个时间点的变形
    x_former = x2_real[:, 2, 2]  # 取前一个时间点的变形
    H=x2_real[0, 0, 0]*x2_real[0, 0, 1]/(2*1.732)
    batch_size = x1.shape[0]
    l1_total=0
    l2_total=0
    l3_total=0
    for i in range(batch_size):
        pi=p[i,0]
        p_formeri=p_former[i,0]
        si=s[i]
        xi=x[i]
        x_formeri=x_former[i]

        if si>0:  # 根据速率判断周丽上升或下降
            l1 = torch.max(torch.tensor(0).to(p_formeri.device), p_formeri - pi)
        else:
            l1=torch.max(torch.tensor(0).to(p_formeri.device), pi - p_formeri)
        l1_total=l1_total+l1**2
        
        if xi*x_formeri<0:  # 根据变形判断临界点
            l2=torch.max(torch.tensor(0).to(p_formeri.device), (p_formeri - 1)*(pi - 1))
        else:
            l2=0
        l2_total=l2_total+l2**2
        
        if xi<-H*10 and si>-10:  # 根据变形速率判断屈曲后承载力（90%可靠度）
            if si<-5:
                l3=torch.max(torch.tensor(0).to(p_formeri.device), pi - 0.55)
            elif si<-1:
                l3=torch.max(torch.tensor(0).to(p_formeri.device), pi + 0.0525*si-0.2875)
            elif si<-0.1:
                l3=torch.max(torch.tensor(0).to(p_formeri.device), pi + 0.211*si-0.129)
            else:
                l3=0
        else:
            l3=0
        l3_total=l3_total+l3**2

    l1_total=l1_total/batch_size
    l2_total=l2_total/batch_size
    l3_total=l3_total/batch_size
    return l1_total, l2_total, l3_total

if __name__ == '__main__':
    train_data_ratio = 0.8  # Choose 80% of the data for training
    max_epochs=10000  #训练循环次数
    batch_size=64
    batch_size_test=64
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    current_device = torch.cuda.current_device()

# ================================ 数据导入和处理
    data_x1 = np.load("data_temperature.npy")
    data_x2 = np.load("data_deformation.npy")
    data_x3 = np.load("data_force.npy")
    data_x2_real = np.load("data_deformation_real.npy")   #未归一化的x2数据，用于物理增强

# ==========================================数据库处理
    np.random.seed(10)
    np.random.shuffle(data_x1)
    np.random.seed(10)
    np.random.shuffle(data_x2)    
    np.random.seed(10)
    np.random.shuffle(data_x3)
    np.random.seed(10)
    np.random.shuffle(data_x2_real)  # 将数据组打乱，为测试集和试验集的选取做准备

    total_batch_size, seq_len = data_x1.shape[0],data_x1.shape[1]
    train_data_len = int(total_batch_size * train_data_ratio)

    train_x1 = data_x1[0:train_data_len]  # 取train_data_len个数据做训练集
    train_x2 = data_x2[0:train_data_len]
    train_x3 = data_x3[0:train_data_len]
    train_x2_real = data_x2_real[0:train_data_len]

    # train数据集需要minibatch
    Mydataset = CustomDataset(train_x1, train_x2, train_x3,train_x2_real)
    train_loader = DataLoader(dataset=Mydataset, batch_size=batch_size, shuffle=False, num_workers=1)

    test_x1 = data_x1[train_data_len:]  # 取剩下数据做测试集
    test_x2 = data_x2[train_data_len:]
    test_x3 = data_x3[train_data_len:]
    test_x2_real = data_x2_real[train_data_len:]
    
    Mydataset_test = CustomDataset(test_x1, test_x2, test_x3,test_x2_real)
    test_loader = DataLoader(dataset=Mydataset_test, batch_size=batch_size_test, shuffle=False, num_workers=0)

# ================================计算pelstm模型
    pelstm_model = PELSTM().to(device)
    pelstm_model=pelstm_model.float()
    optimizer = torch.optim.Adam(pelstm_model.parameters(), lr=1e-5)

    prev_loss = 1000
    loss_value_data = []
    loss_value_physics1 = []
    loss_value_physics2 = []
    loss_value_physics3 = []
    test_loss_value = []
    for epoch in range(max_epochs):
        for i, data in enumerate(train_loader, 0):
            train_x1_tensor, train_x2_tensor, train_x3_tensor, train_x2_real_tensor = data
            train_x1_tensor=train_x1_tensor.float().to(device)
            train_x2_tensor=train_x2_tensor.float().to(device)
            train_x3_tensor=train_x3_tensor.float().to(device)
            train_x2_real_tensor=train_x2_real_tensor.float().to(device)
            optimizer.zero_grad()
            loss_data = data_loss(pelstm_model, train_x1_tensor, train_x2_tensor, train_x3_tensor)
            loss = loss_data
            loss.backward()
            clip_grad_norm_(pelstm_model.parameters(), max_norm=0.1)
            optimizer.step()
            torch.cuda.empty_cache()
            if loss < prev_loss:
                torch.save(pelstm_model.state_dict(), 'lstm_64batch.pt')  # save model parameters to files
                prev_loss = loss
                print('Epoch: [{}/{}], Loss:{:.5f}'.format(epoch + 1, max_epochs, loss.item()))
            loss_data_w=np.expand_dims(loss_data.item(),axis=0)
            loss_value_data.extend(loss_data_w)         
            
        for k, data_test in enumerate(test_loader, 0):
            test_x1_tensor, test_x2_tensor, test_x3_tensor, test_x2_real_tensor = data_test
            test_x1_tensor=test_x1_tensor.float().to(device)
            test_x2_tensor=test_x2_tensor.float().to(device)
            test_x3_tensor=test_x3_tensor.float().to(device)
            test_x2_real_tensor=test_x2_real_tensor.float().to(device)
            loss_test_data = data_loss(pelstm_model, test_x1_tensor, test_x2_tensor, test_x3_tensor)
            loss_test = loss_test_data
            torch.cuda.empty_cache()
            test_loss_value_w=np.expand_dims(loss_test.item(),axis=0)
            test_loss_value.extend(test_loss_value_w)

        if loss.item() < 1e-7:
            print('Epoch [{}/{}], Loss: {:.5f}'.format(epoch + 1, max_epochs, loss.item()))
            print("The loss value is reached")
            break

    trainbatch_geshu=i+1  
    testbatch_geshu=k+1


# ===============================输出loss
    loss_value_data=np.array(loss_value_data).reshape(-1,1)
    loss_value_physics1=np.array(loss_value_physics1).reshape(-1,1)
    loss_value_physics2=np.array(loss_value_physics2).reshape(-1,1)
    loss_value_physics3=np.array(loss_value_physics3).reshape(-1,1)
    loss_geshu=loss_value_data.shape[0]
    test_loss_value=np.array(test_loss_value).reshape(-1,1)
    test_loss_geshu=test_loss_value.shape[0]
    wb = Workbook()
    ws1 = wb.active
    ws2=wb.create_sheet()
    ws3=wb.create_sheet()
    ws4=wb.create_sheet()
    ws5=wb.create_sheet()
    ws1.title = "train loss_data"
    row=0
    for i in range(1, loss_geshu+1,trainbatch_geshu):
        row = row + 1
        for j in range(1,trainbatch_geshu+1):
            ws1.cell(row=row, column=j).value = loss_value_data[i + j - 2, 0]         
    ws5.title = "test loss"
    row=0
    for i in range(1, test_loss_geshu+1,testbatch_geshu):
        row = row + 1
        for j in range(1,testbatch_geshu+1):
            ws5.cell(row=row, column=j).value = test_loss_value[i + j - 2, 0]
    wb.save('loss_lstm_64batch.xlsx')